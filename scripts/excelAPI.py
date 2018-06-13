import openpyxl
import xlrd
import os
from openpyxl import Workbook
from scripts.weather import weather
from xlutils.copy import copy

ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
# filename = "./Excelsheets/AIML.xls"
workbookpath=ROOT_DIR.replace("scripts","Excelsheets")
wsToOpen = "Weather"
workbookold = openpyxl.load_workbook(os.path.join(workbookpath,'AIML.xlsx'),data_only=True)
# workbook=copy(workbookold)
worksheet = workbookold['Weather']
print(worksheet.cell(1, 1).value)
for i in range(1,1000):
        ZIPCODE=worksheet.cell(i+1,1).value
        weather(ZIPCODE)
        mainweatherReport=weather(ZIPCODE)
        worksheet.cell(i+1,2).value=mainweatherReport
        savepath=workbookpath+"\\"+'AIML.xlsx'
        workbookold.save(savepath)
        # workbookold.save()
        if(worksheet.cell(i+1,1).value==None):
            break
print("Done")


